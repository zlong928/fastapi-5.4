"""
Excel导出器：将图表提取结果输出为格式化的Excel文件
支持所有图表类型：坐标图、条形图、热图、表格图等
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ChartExcelExporter:
    """通用图表数据导出器（支持所有图表类型）"""

    def __init__(self, output_path: str | Path | None = None) -> None:
        if not OPENPYXL_AVAILABLE:
            raise ImportError("需要安装openpyxl: pip install openpyxl")
        self.output_path = Path(output_path) if output_path else None
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def export(self, extraction_results: dict, pdf_filename: str = "document") -> str | bytes:
        """导出所有图表数据到Excel

        Args:
            extraction_results: 包含 by_figure 的提取结果字典
            pdf_filename: 源文档名称

        Returns:
            如果指定了 output_path 返回路径字符串，否则返回字节流
        """
        # 创建数据工作表
        ws_data = self.wb.create_sheet("Chart Data")
        ws_long = self.wb.create_sheet("Long Data")
        current_row = 1
        long_row = 1

        figures_extracted = []
        low_confidence = []

        # 写入 Long Data 表头
        ws_long.cell(long_row, 1, "Figure")
        ws_long.cell(long_row, 2, "Panel")
        ws_long.cell(long_row, 3, "Chart Type")
        ws_long.cell(long_row, 4, "Series/Category")
        ws_long.cell(long_row, 5, "X/Category")
        ws_long.cell(long_row, 6, "Y/Value")
        ws_long.cell(long_row, 7, "Error/Extra")
        for col in range(1, 8):
            ws_long.cell(long_row, col).font = Font(bold=True)
        long_row += 1

        # 遍历所有图表结果
        for figure_id, visual_result in extraction_results.get("by_figure", {}).items():
            chart_data = visual_result.get("chart_data") or visual_result.get("coordinate_data")
            if not chart_data or "error" in chart_data:
                continue

            chart_type = chart_data.get("chart_type", "unknown")
            confidence = chart_data.get("confidence", 1.0)

            figures_extracted.append(f"{figure_id} ({chart_type})")

            if confidence < 0.6:
                low_confidence.append({
                    "figure": figure_id,
                    "chart_type": chart_type,
                    "confidence": confidence,
                    "reason": chart_data.get("unsupported_reason", "Low confidence")
                })

            # 根据图表类型路由
            if chart_type in (
                "line_plot",
                "biphasic_time_series",
                "multi_line_plot",
                "rheology_flow_curve",
                "rheology_strain_sweep",
                "rheology_step_time_sweep",
                "scatter_plot",
                "spectrum_curve",
                "bar_or_line_with_errorbar",
                "generic_coordinate_plot",
                "dual_axis_plot",
            ):
                current_row, long_row = self._write_coordinate_chart(ws_data, ws_long, chart_data, figure_id, current_row, long_row)
            elif chart_type in ("bar_chart", "grouped_bar"):
                current_row, long_row = self._write_bar_chart(ws_data, ws_long, chart_data, figure_id, current_row, long_row)
            elif chart_type in ("heatmap", "heatmap_matrix", "2d_field_map"):
                current_row, long_row = self._write_heatmap(ws_data, ws_long, chart_data, figure_id, current_row, long_row)
            elif chart_type == "table_image":
                current_row, long_row = self._write_table_image(ws_data, ws_long, chart_data, figure_id, current_row, long_row)
            else:
                # 通用处理
                current_row, long_row = self._write_generic_chart(ws_data, ws_long, chart_data, figure_id, current_row, long_row)

            # 块之间空2行
            current_row += 2

        # 创建Notes和Low Confidence工作表
        self._write_notes_sheet(pdf_filename, figures_extracted)
        if low_confidence:
            self._write_low_confidence_sheet(low_confidence)

        # 保存或返回字节流
        if self.output_path:
            self.wb.save(self.output_path)
            return str(self.output_path)
        else:
            buffer = BytesIO()
            self.wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

    def _write_coordinate_chart(self, ws: Any, ws_long: Any, chart_data: dict, figure_id: str, start_row: int, long_row: int) -> tuple[int, int]:
        """写入坐标图数据"""
        panels = chart_data.get("panels", [])
        if not panels:
            return start_row, long_row

        for panel_data in panels:
            panel = panel_data.get("panel", "")
            x_axis = panel_data.get("x_axis", {})
            y_axis = panel_data.get("y_axis", {})
            method = panel_data.get("extraction_method", "unknown")

            # 标题行
            title = f"=== {figure_id} Panel {panel}: coordinate_plot ==="
            ws.cell(start_row, 1, title)
            ws.cell(start_row, 1).font = Font(bold=True, size=11)
            ws.cell(start_row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            start_row += 1

            # 元信息
            ws.cell(start_row, 1, f"Method: {method} | X: {x_axis.get('label', '')} ({x_axis.get('unit', '')}) | Y: {y_axis.get('label', '')} ({y_axis.get('unit', '')})")
            start_row += 1

            # 数据表头
            series_list = panel_data.get("series", [])
            ws.cell(start_row, 1, "X")
            for idx, series in enumerate(series_list):
                ws.cell(start_row, idx + 2, series.get("series_name", f"Series{idx+1}"))
                ws.cell(start_row, idx + 2).font = Font(bold=True)
            ws.cell(start_row, 1).font = Font(bold=True)
            start_row += 1

            # 收集所有 x 值
            all_x = set()
            for series in series_list:
                for point in series.get("data_points", []):
                    all_x.add(point.get("x"))
            all_x = sorted([x for x in all_x if x is not None])

            # 写入数据行
            for x_val in all_x:
                ws.cell(start_row, 1, x_val)
                for idx, series in enumerate(series_list):
                    # 查找对应的 y 值
                    y_val = None
                    for point in series.get("data_points", []):
                        if point.get("x") == x_val:
                            y_val = point.get("y")
                            break
                    ws.cell(start_row, idx + 2, y_val if y_val is not None else "/")

                    # 写入 Long Data
                    if y_val is not None:
                        ws_long.cell(long_row, 1, figure_id)
                        ws_long.cell(long_row, 2, panel)
                        ws_long.cell(long_row, 3, "coordinate_plot")
                        ws_long.cell(long_row, 4, series.get("series_name", ""))
                        ws_long.cell(long_row, 5, x_val)
                        ws_long.cell(long_row, 6, y_val)
                        long_row += 1
                start_row += 1

        return start_row, long_row

    def _write_bar_chart(self, ws: Any, ws_long: Any, chart_data: dict, figure_id: str, start_row: int, long_row: int) -> tuple[int, int]:
        """写入条形图数据"""
        panel = chart_data.get("panel", "")
        x_axis = chart_data.get("x_axis", {})
        y_axis = chart_data.get("y_axis", {})
        method = chart_data.get("extraction_method", "unknown")

        # 标题行
        title = f"=== {figure_id} Panel {panel}: bar_chart ==="
        ws.cell(start_row, 1, title)
        ws.cell(start_row, 1).font = Font(bold=True, size=11)
        ws.cell(start_row, 1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        start_row += 1

        # 元信息
        ws.cell(start_row, 1, f"Method: {method} | Y: {y_axis.get('label', '')} ({y_axis.get('unit', '')})")
        start_row += 1

        # 数据表头
        series_list = chart_data.get("series", [])
        ws.cell(start_row, 1, "Category")
        for idx, series in enumerate(series_list):
            ws.cell(start_row, idx + 2, series.get("series_name", f"Series{idx+1}"))
            ws.cell(start_row, idx + 2).font = Font(bold=True)
        ws.cell(start_row, 1).font = Font(bold=True)
        start_row += 1

        # 收集所有类别
        all_categories = []
        if series_list:
            for bar in series_list[0].get("bars", []):
                cat = bar.get("category")
                if cat and cat not in all_categories:
                    all_categories.append(cat)

        # 写入数据行
        for category in all_categories:
            ws.cell(start_row, 1, category)
            for idx, series in enumerate(series_list):
                # 查找对应的值
                value = None
                error_bar = None
                for bar in series.get("bars", []):
                    if bar.get("category") == category:
                        value = bar.get("value")
                        error_bar = bar.get("error_bar")
                        break
                display_value = f"{value}" if value is not None else "/"
                if error_bar:
                    display_value += f" ± {error_bar}"
                ws.cell(start_row, idx + 2, display_value)

                # 写入 Long Data
                if value is not None:
                    ws_long.cell(long_row, 1, figure_id)
                    ws_long.cell(long_row, 2, panel)
                    ws_long.cell(long_row, 3, "bar_chart")
                    ws_long.cell(long_row, 4, series.get("series_name", ""))
                    ws_long.cell(long_row, 5, category)
                    ws_long.cell(long_row, 6, value)
                    ws_long.cell(long_row, 7, error_bar if error_bar else "")
                    long_row += 1
            start_row += 1

        return start_row, long_row

    def _write_heatmap(self, ws: Any, ws_long: Any, chart_data: dict, figure_id: str, start_row: int, long_row: int) -> tuple[int, int]:
        """写入热图数据"""
        panel = chart_data.get("panel", "")
        x_axis = chart_data.get("x_axis", {})
        y_axis = chart_data.get("y_axis", {})
        method = chart_data.get("extraction_method", "llm_assisted_reading")

        # 标题行
        title = f"=== {figure_id} Panel {panel}: heatmap ==="
        ws.cell(start_row, 1, title)
        ws.cell(start_row, 1).font = Font(bold=True, size=11)
        ws.cell(start_row, 1).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        start_row += 1

        # 元信息
        ws.cell(start_row, 1, f"Method: {method} | Note: Values may be approximate")
        start_row += 1

        # 获取矩阵和轴标签
        matrix = chart_data.get("matrix", [])
        x_values = x_axis.get("values", [])
        y_categories = y_axis.get("categories", [])

        # 表头
        ws.cell(start_row, 1, "")
        for idx, x_val in enumerate(x_values):
            ws.cell(start_row, idx + 2, x_val)
            ws.cell(start_row, idx + 2).font = Font(bold=True)
        start_row += 1

        # 数据行
        for row_idx, (y_cat, row_data) in enumerate(zip(y_categories, matrix)):
            ws.cell(start_row, 1, y_cat)
            ws.cell(start_row, 1).font = Font(bold=True)
            for col_idx, value in enumerate(row_data):
                ws.cell(start_row, col_idx + 2, value if value is not None else "/")

                # 写入 Long Data
                if value is not None and col_idx < len(x_values):
                    ws_long.cell(long_row, 1, figure_id)
                    ws_long.cell(long_row, 2, panel)
                    ws_long.cell(long_row, 3, "heatmap")
                    ws_long.cell(long_row, 4, y_cat)
                    ws_long.cell(long_row, 5, x_values[col_idx])
                    ws_long.cell(long_row, 6, value)
                    long_row += 1
            start_row += 1

        return start_row, long_row

    def _write_table_image(self, ws: Any, ws_long: Any, chart_data: dict, figure_id: str, start_row: int, long_row: int) -> tuple[int, int]:
        """写入表格图数据"""
        # 标题行
        title = f"=== {figure_id}: table_image ==="
        ws.cell(start_row, 1, title)
        ws.cell(start_row, 1).font = Font(bold=True, size=11)
        ws.cell(start_row, 1).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        start_row += 1

        table_data = chart_data.get("table_data", {})
        headers = table_data.get("headers", [])
        rows = table_data.get("rows", [])

        # 表头
        for idx, header in enumerate(headers):
            ws.cell(start_row, idx + 1, header)
            ws.cell(start_row, idx + 1).font = Font(bold=True)
        start_row += 1

        # 数据行
        for row in rows:
            for idx, cell_value in enumerate(row):
                ws.cell(start_row, idx + 1, cell_value)

                # 写入 Long Data
                if idx < len(headers):
                    ws_long.cell(long_row, 1, figure_id)
                    ws_long.cell(long_row, 2, "")
                    ws_long.cell(long_row, 3, "table_image")
                    ws_long.cell(long_row, 4, headers[idx] if idx < len(headers) else "")
                    ws_long.cell(long_row, 5, row[0] if row else "")
                    ws_long.cell(long_row, 6, cell_value)
                    long_row += 1
            start_row += 1

        return start_row, long_row

    def _write_generic_chart(self, ws: Any, ws_long: Any, chart_data: dict, figure_id: str, start_row: int, long_row: int) -> tuple[int, int]:
        """写入通用图表数据（不支持的类型）"""
        chart_type = chart_data.get("chart_type", "unknown")
        panel = chart_data.get("panel", "")

        # 标题行
        title = f"=== {figure_id} Panel {panel}: {chart_type} ==="
        ws.cell(start_row, 1, title)
        ws.cell(start_row, 1).font = Font(bold=True, size=11)
        start_row += 1

        description = chart_data.get("description", "")
        if description:
            ws.cell(start_row, 1, f"Description: {description}")
            start_row += 1

        unsupported_reason = chart_data.get("unsupported_reason", "")
        if unsupported_reason:
            ws.cell(start_row, 1, f"Note: {unsupported_reason}")
            start_row += 1

        return start_row, long_row

    def _write_notes_sheet(self, pdf_filename: str, figures_extracted: list[str]) -> None:
        """创建Notes工作表"""
        ws = self.wb.create_sheet("Notes")
        row = 1

        ws.cell(row, 1, "Chart Data Extraction Summary")
        ws.cell(row, 1).font = Font(bold=True, size=14)
        row += 2

        ws.cell(row, 1, f"Source Document: {pdf_filename}")
        row += 1

        ws.cell(row, 1, f"Total Figures Extracted: {len(figures_extracted)}")
        row += 2

        ws.cell(row, 1, "Figures:")
        ws.cell(row, 1).font = Font(bold=True)
        row += 1

        for fig in figures_extracted:
            ws.cell(row, 1, f"  - {fig}")
            row += 1

    def _write_low_confidence_sheet(self, low_confidence: list[dict]) -> None:
        """创建 Low Confidence 工作表"""
        ws = self.wb.create_sheet("Low Confidence")
        row = 1

        # 表头
        ws.cell(row, 1, "Figure")
        ws.cell(row, 2, "Chart Type")
        ws.cell(row, 3, "Confidence")
        ws.cell(row, 4, "Reason")
        for col in range(1, 5):
            ws.cell(row, col).font = Font(bold=True)
        row += 1

        # 数据行
        for item in low_confidence:
            ws.cell(row, 1, item["figure"])
            ws.cell(row, 2, item["chart_type"])
            ws.cell(row, 3, item["confidence"])
            ws.cell(row, 4, item["reason"])
            row += 1


# 向后兼容别名
class CoordinateExcelExporter(ChartExcelExporter):
    """向后兼容的坐标导出器别名"""
    pass


def export_from_results(extraction_results: dict, filename: str = "extraction") -> bytes:
    """从提取结果直接生成 Excel 字节流

    Args:
        extraction_results: 包含 by_figure 的提取结果字典
        filename: 文档名称（用于 Notes）

    Returns:
        Excel 文件的字节流
    """
    exporter = ChartExcelExporter()
    return exporter.export(extraction_results, filename)
